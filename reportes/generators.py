# reportes/generators.py
import datetime
from decimal import Decimal
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

def _limpiar_valor(valor):
    """Convierte valores especiales (Decimal, Fecha, None) a strings legibles."""
    if isinstance(valor, Decimal):
        return f"{valor:,.2f}"
    if isinstance(valor, (datetime.date, datetime.datetime)):
        return valor.strftime("%Y-%m-%d")
    if valor is None:
        return ""
    return str(valor)

# ===================================================================
# --- GENERADOR DE REPORTE EXCEL (OPENPYXL) ---
# ===================================================================
def generar_reporte_excel(data, interpretacion):
    """
    Genera un archivo Excel (xlsx) en memoria a partir de una lista de diccionarios.
    """
    prompt_titulo = interpretacion.get('prompt', 'Reporte')
    
    # --- Configuración de la Respuesta HTTP ---
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.date.today()}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"

    if not data:
        ws['A1'] = "No se encontraron datos para este reporte."
        wb.save(response)
        return response

    # --- Estilos ---
    font_titulo = Font(bold=True, size=14)
    font_header = Font(bold=True, color="FFFFFF")
    fill_header = colors.Color(rgb='004A99') # Un azul corporativo
    alignment_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # --- Título ---
    ws['A1'] = prompt_titulo
    ws.merge_cells('A1:F1')
    ws['A1'].font = font_titulo
    ws['A1'].alignment = Alignment(horizontal='left')

    # --- Encabezados (Headers) ---
    headers = list(data[0].keys())
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header_title.replace("_", " ").title())
        cell.font = font_header
        cell.fill = colors.PatternFill(start_color=fill_header, end_color=fill_header, fill_type="solid")
        cell.alignment = alignment_center
        cell.border = border_thin

    # --- Escribir Datos ---
    row_num = 4
    for row_data in data:
        for col_num, header in enumerate(headers, 1):
            valor = row_data.get(header)
            cell = ws.cell(row=row_num, column=col_num, value=_limpiar_valor(valor))
            cell.border = border_thin
        row_num += 1

    # --- Ajustar Ancho de Columnas (Auto-fit) ---
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].autosize = True

    # Guardar y devolver
    wb.save(response)
    return response

# ===================================================================
# --- GENERADOR DE REPORTE PDF (REPORTLAB) ---
# ===================================================================
def generar_reporte_pdf(data, interpretacion):
    """
    Genera un archivo PDF en memoria a partir de una lista de diccionarios.
    """
    prompt_titulo = interpretacion.get('prompt', 'Reporte')

    # --- Configuración de la Respuesta HTTP ---
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{datetime.date.today()}.pdf"'

    # --- Configuración del Documento ---
    # Usamos landscape (horizontal) para que quepan más columnas
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    story = []
    styles = getSampleStyleSheet()

    if not data:
        story.append(Paragraph("No se encontraron datos para este reporte.", styles['Normal']))
        doc.build(story)
        return response

    # --- Título ---
    story.append(Paragraph(prompt_titulo, styles['h1']))
    story.append(Paragraph(f"Generado el: {datetime.date.today()}", styles['Normal']))
    story.append(Spacer(1, 0.25*inch))

    # --- Preparar Datos de la Tabla ---
    headers = list(data[0].keys())
    # Limpiamos los headers para mostrarlos
    clean_headers = [h.replace("_", " ").title() for h in headers]
    
    # Convertimos todos los datos a strings usando el helper
    table_data = [clean_headers] + [
        [_limpiar_valor(row.get(header)) for header in headers] for row in data
    ]

    # --- Crear Tabla ---
    t = Table(table_data, repeatRows=1) # Repetir headers en cada página

    # --- Estilo de la Tabla ---
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#004A99')), # Header azul
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # Header en negrita
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('FONTSIZE', (0, 1), (-1, -1), 8), # Fuente de datos más pequeña
        ('GRID', (0, 0), (-1, -1), 1, colors.black), # Bordes
    ])
    
    t.setStyle(style)
    story.append(t)

    # Construir y devolver
    doc.build(story)
    return response